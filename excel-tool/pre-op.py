import os
from openpyxl import Workbook
from openpyxl import load_workbook


src_file_path = 'C:/Users/wly-PC/Desktop/main_src.xlsx'
dst_file_path = os.path.splitext(src_file_path)[0] + '_mod.xlsx'
wb = load_workbook(src_file_path)
# create a new workbook
# wb = Workbook()

# grab the active worksheet
ws = wb.active

# for cols in ws.iter_rows(min_row=2,  max_col=1):
# 	for cell in cols:
# 		# print(cell.coordinate)
# 		cell.value = cell.value.replace(' ', '')
# 		print(cell.value)


for row in ws.iter_rows(min_row=2, max_col=1):
	print(row[0])
	#ws.row_dimensions[row].height = 15
	for cell in row:
		# print(cell.coordinate)
		cell.value = cell.value.replace(' ', '')
		#print(cell.value)


# Data can be assigned directly to cells
# ws['A1'] = 42

# Rows can also be appended
#ws.append([1, 2, 3])

# Python types will automatically be converted
# import datetime
# ws['A2'] = datetime.datetime.now()

# Save the file
wb.save(dst_file_path)